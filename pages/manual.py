import re
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from openpyxl import load_workbook, Workbook

from models import Client
from extensions import db

bp = Blueprint("manual", __name__, url_prefix="/manual")


def _only_digits(s: str) -> str:
    return re.sub(r"\D", "", s or "")


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e")
    s = s.replace("í", "i")
    s = s.replace("õ", "o").replace("ó", "o").replace("ô", "o")
    s = s.replace("ç", "c")
    s = re.sub(r"\s+", "_", s)
    return s


@bp.get("/")
@login_required
def index():
    q = (request.args.get("q") or "").strip()
    regime = (request.args.get("regime") or "").strip()

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    query = Client.query

    if q:
        q_like = f"%{q}%"
        query = query.filter(
            (Client.razao_social.ilike(q_like))
            | (Client.cnpj.ilike(q_like))
            | (Client.responsavel_fiscal.ilike(q_like))
        )

    if regime:
        regime_like = f"%{regime}%"
        query = query.filter(Client.regime_tributario.ilike(regime_like))

    pagination = query.order_by(Client.razao_social.asc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )

    return render_template(
        "manual/index.html",
        clients=pagination.items,
        pagination=pagination,
        q=q,
        regime=regime,
        per_page=per_page
    )


@bp.route("/novo", methods=["GET", "POST"])
@login_required
def new_client():
    if request.method == "POST":
        razao_social = (request.form.get("razao_social") or "").strip()
        cnpj_raw = (request.form.get("cnpj") or "").strip()
        regime_tributario = (request.form.get("regime_tributario") or "").strip()
        responsavel_fiscal = (request.form.get("responsavel_fiscal") or "").strip()

        if not (razao_social and cnpj_raw and regime_tributario and responsavel_fiscal):
            flash("Preencha todos os campos.", "warning")
            return redirect(url_for("manual.new_client"))

        cnpj = _only_digits(cnpj_raw)

        if len(cnpj) != 14:
            flash("CNPJ inválido. Digite um CNPJ com 14 números.", "warning")
            return redirect(url_for("manual.new_client"))

        exists = Client.query.filter_by(cnpj=cnpj).first()
        if exists:
            flash("Já existe um cliente cadastrado com esse CNPJ.", "warning")
            return redirect(url_for("manual.new_client"))

        client = Client(
            razao_social=razao_social,
            cnpj=cnpj,
            regime_tributario=regime_tributario,
            responsavel_fiscal=responsavel_fiscal,
        )
        db.session.add(client)
        db.session.commit()

        flash("Cliente cadastrado com sucesso!", "success")
        return redirect(url_for("manual.index"))

    return render_template("manual/new_client.html")


# ✅ ROTA ÚNICA: gera e baixa um Excel modelo válido (sem depender de arquivo físico)
@bp.get("/modelo-clientes")
@login_required
def modelo_clientes():
    wb = Workbook()
    ws = wb.active
    ws.title = "clientes"

    # cabeçalho (colunas que seu import entende)
    ws.append(["razao_social", "cnpj", "regime_tributario", "responsavel_fiscal"])

    # linha exemplo (opcional)
    ws.append(["EMPRESA EXEMPLO LTDA", "00.000.000/0000-00", "SIMPLES NACIONAL", "RESPONSÁVEL"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="modelo_clientes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ✅ NOVA ROTA: exporta os clientes do banco para Excel (respeitando q e regime)
@bp.get("/exportar")
@login_required
def export_xlsx():
    q = (request.args.get("q") or "").strip()
    regime = (request.args.get("regime") or "").strip()

    query = Client.query

    if q:
        q_like = f"%{q}%"
        query = query.filter(
            (Client.razao_social.ilike(q_like))
            | (Client.cnpj.ilike(q_like))
            | (Client.responsavel_fiscal.ilike(q_like))
        )

    if regime:
        regime_like = f"%{regime}%"
        query = query.filter(Client.regime_tributario.ilike(regime_like))

    clients = query.order_by(Client.razao_social.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "clientes"

    ws.append(["codigo", "razao_social", "cnpj", "regime_tributario", "responsavel_fiscal"])

    for c in clients:
        ws.append([
            c.id,
            c.razao_social or "",
            str(c.cnpj or ""),  # força texto (CNPJ não vira número)
            c.regime_tributario or "",
            c.responsavel_fiscal or "",
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="clientes_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/importar", methods=["GET", "POST"])
@login_required
def import_xlsx():
    if request.method == "GET":
        return render_template("manual/import_xlsx.html")

    file = request.files.get("file")
    atualizar = (request.form.get("atualizar") == "1")

    if not file or not file.filename:
        flash("Selecione um arquivo .xlsx para importar.", "warning")
        return redirect(url_for("manual.import_xlsx"))

    if not (file.filename or "").lower().endswith(".xlsx"):
        flash("Arquivo inválido. Envie um arquivo .xlsx.", "warning")
        return redirect(url_for("manual.import_xlsx"))

    try:
        wb = load_workbook(file.stream, data_only=True, read_only=True)
        ws = wb.active
    except Exception:
        flash("Não consegui ler esse Excel. Confirme se é .xlsx válido.", "danger")
        return redirect(url_for("manual.import_xlsx"))

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        flash("Planilha vazia.", "warning")
        return redirect(url_for("manual.import_xlsx"))

    header = [_norm_header(str(c or "")) for c in rows[0]]

    aliases = {
        "razao_social": {"razao_social", "razao", "razao_social_empresa", "nome", "empresa", "cliente"},
        "cnpj": {"cnpj", "cnpj_cpf", "documento", "documento_de_identificacao"},
        "regime_tributario": {"regime_tributario", "regime", "regime_federal", "tributacao"},
        "responsavel_fiscal": {"responsavel_fiscal", "responsavel", "resp_fiscal", "responsavel_do_fiscal"},
    }

    col = {}
    for idx, h in enumerate(header):
        for key, opts in aliases.items():
            if h in opts and key not in col:
                col[key] = idx

    missing = [k for k in aliases.keys() if k not in col]
    if missing:
        flash("Colunas obrigatórias não encontradas: " + ", ".join(missing), "danger")
        return redirect(url_for("manual.import_xlsx"))

    inserted = 0
    updated = 0
    skipped = 0
    errors = 0

    try:
        for r in rows[1:]:
            try:
                razao_social = str(r[col["razao_social"]] or "").strip()
                cnpj_raw = str(r[col["cnpj"]] or "").strip()
                regime_tributario = str(r[col["regime_tributario"]] or "").strip()
                responsavel_fiscal = str(r[col["responsavel_fiscal"]] or "").strip()

                if not (razao_social and cnpj_raw and regime_tributario and responsavel_fiscal):
                    errors += 1
                    continue

                cnpj = _only_digits(cnpj_raw)
                if len(cnpj) != 14:
                    errors += 1
                    continue

                exists = Client.query.filter_by(cnpj=cnpj).first()
                if exists:
                    if atualizar:
                        exists.razao_social = razao_social
                        exists.regime_tributario = regime_tributario
                        exists.responsavel_fiscal = responsavel_fiscal
                        updated += 1
                    else:
                        skipped += 1
                    continue

                client = Client(
                    razao_social=razao_social,
                    cnpj=cnpj,
                    regime_tributario=regime_tributario,
                    responsavel_fiscal=responsavel_fiscal,
                )
                db.session.add(client)
                inserted += 1

            except Exception:
                errors += 1
                continue

        db.session.commit()

    except Exception:
        db.session.rollback()
        flash("Falha ao importar: erro ao gravar no banco de dados.", "danger")
        return redirect(url_for("manual.import_xlsx"))

    flash(
        f"Importação concluída. Inseridos: {inserted} | Atualizados: {updated} | "
        f"Duplicados ignorados: {skipped} | Linhas com erro: {errors}",
        "success" if (inserted or updated) else "warning",
    )
    return redirect(url_for("manual.index"))


@bp.get("/cliente/<int:client_id>")
@login_required
def client_detail(client_id: int):
    client = Client.query.get_or_404(client_id)
    return render_template("manual/client_detail.html", client=client)


# --- Fallback de segurança: garante que manual.index exista ---
if "index" not in bp.view_functions:
    bp.add_url_rule("/", endpoint="index", view_func=index, methods=["GET"])
